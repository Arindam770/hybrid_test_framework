
import openpyxl as xl


class Excel_Helper:

    def __init__(self, path):
        self.path = path

    def read_excel(self):
        workbook = xl.load_workbook(self.path)
        sheet = workbook.active

        header = []
        for cell in sheet[1]:
            header.append(cell.value)

        all_excel_details = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            temp_dict = {}
            for num in range(len(header)):
                if row[num] is not None:
                    temp_dict[header[num]] = row[num]
                else:
                    temp_dict[header[num]] = ""
            all_excel_details.append(temp_dict)
        print(all_excel_details)
        return all_excel_details

    def write_excel(self, data):
        header = list(data[0].keys())

        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.append(header)

        for row in data:
            temp_list = []
            for elm in header:
                temp_list.append(row.get(elm,""))
            sheet.append(temp_list)
        workbook.save(self.path)

if __name__ == "__main__":
    excel_obj = Excel_Helper("sample_file\\DummyData.xlsx")
    data = excel_obj.read_excel()
    excel_obj.write_excel(data)